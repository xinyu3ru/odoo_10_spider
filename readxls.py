#!/bin/env python
# -*- coding: utf-8 -*-

__author__ = 'rublog'

import openpyxl
import logging

logger = logging.getLogger()
logger.setLevel('DEBUG')
BASIC_FORMAT = "%(asctime)s:%(levelname)s:%(message)s"
DATE_FORMAT = '%Y-%m-%d %H:%M:%S'
formatter = logging.Formatter(BASIC_FORMAT, DATE_FORMAT)
cons_log = logging.StreamHandler()  # 输出到控制台的handler
cons_log.setFormatter(formatter)
cons_log.setLevel('INFO')  # 也可以不设置，不设置就默认用logger的level
file_log = logging.FileHandler('output.log')  # 输出到文件的handler
file_log.setFormatter(formatter)
logger.addHandler(cons_log)
logger.addHandler(file_log)
logger.info('this is info')
logger.debug('this is debug')


def read_xls(xls_name='2020.xlsx', sheet_name="rublog"):
    wb = openpyxl.load_workbook(xls_name)
    logger.info(wb.sheetnames)
    wx = wb[sheet_name]
    col_a = wx['A4'].value
    logger.info(col_a)
    for i in range(3, wx.max_row):
        col_num = "A" + str(i)
        col_a = wx[col_num].value
        if col_a[0:2] in ["石家", "唐山", "秦皇", "邯郸", "邢台", "保定", "张家", "承德", "沧州", "廊坊", "太原", "衡水", "大同", "阳泉", "长治", "晋城", "朔州", "晋中", "运城", "忻州", "临汾", "吕梁", "呼和", "包头", "乌海", "赤峰", "通辽", "鄂尔", "呼伦", "巴彦", "乌兰", "兴安", "锡林", "阿拉", "沈阳", "大连", "鞍山", "抚顺", "本溪", "丹东", "锦州", "营口", "阜新", "辽阳", "盘锦", "铁岭", "朝阳", "葫芦", "长春", "吉林", "四平", "辽源", "通化", "白山", "松原", "白城", "延边", "鲜族", "哈尔", "齐齐", "鸡西", "鹤岗", "双鸭", "大庆", "伊春", "佳木", "七台", "牡丹", "黑河", "绥化", "大兴", "南京", "无锡", "徐州", "常州", "苏州", "南通", "连云", "淮安", "盐城", "扬州", "镇江", "泰州", "宿迁", "杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华", "衢州", "舟山", "台州", "丽水", "合肥", "芜湖", "蚌埠", "淮南",  "马鞍", "淮北", "铜陵", "安庆", "黄山", "滁州", "阜阳", "宿州", "六安", "亳州", "池州", "宣城", "福州", "厦门", "莆田", "三明", "泉州", "漳州", "南平", "龙岩", "宁德", "南昌", "景德", "萍乡", "九江", "新余", "鹰潭", "赣州", "吉安", "宜春", "抚州", "上饶", "济南", "青岛", "淄博", "枣庄", "东营", "烟台", "潍坊", "济宁", "泰安", "威海", "日照", "莱芜", "临沂", "德州", "聊城", "滨州", "菏泽", "郑州", "开封", "洛阳", "平顶", "安阳", "鹤壁", "新乡", "焦作", "濮阳", "许昌", "漯河", "三门", "南阳", "商丘", "信阳", "周口", "驻马", "济源", "武汉", "黄石", "十堰", "宜昌", "襄阳", "鄂州", "荆门", "孝感", "荆州", "黄冈", "咸宁", "随州", "恩施", "长沙", "株洲", "湘潭", "衡阳", "邵阳", "岳阳", "常德", "张家", "益阳", "郴州", "永州", "怀化", "娄底", "湘西", "广州", "韶关", "深圳", "珠海", "汕头", "佛山", "江门", "湛江", "茂名", "肇庆", "惠州", "梅州", "汕尾", "河源", "阳江", "清远", "东莞", "中山", "潮州", "揭阳", "云浮", "南宁", "柳州", "桂林", "梧州", "北海", "防城", "钦州", "贵港", "玉林", "百色", "贺州", "河池", "来宾", "崇左", "海口", "三亚", "三沙", "儋州", "成都", "自贡", "攀枝", "泸州", "德阳", "绵阳", "广元", "遂宁", "内江", "乐山", "南充", "眉山", "宜宾", "广安", "达州", "雅安", "巴中", "资阳", "阿坝", "甘孜", "凉山", "贵阳", "六盘", "遵义", "安顺", "毕节", "铜仁", "黔西", "黔东", "黔南", "昆明", "曲靖", "玉溪", "保山", "昭通", "丽江", "普洱", "临沧", "楚雄", "红河", "文山", "西双", "大理", "德宏", "怒江", "迪庆", "拉萨", "昌都", "山南", "日喀", "那曲", "阿里", "林芝", "西安", "铜川", "宝鸡", "咸阳", "渭南", "延安", "汉中", "榆林", "安康", "商洛", "兰州", "嘉峪",  "金昌", "白银", "天水", "武威", "张掖", "平凉", "酒泉", "庆阳", "定西", "陇南", "临夏", "甘南", "西宁", "海东", "海北", "黄南", "海南", "果洛", "玉树", "海西", "银川", "石嘴", "吴忠", "固原", "中卫", "乌鲁", "克拉", "吐鲁", "哈密", "昌吉", "博尔", "巴音", "阿克", "克孜", "喀什", "和田", "伊犁", "塔城", "阿勒"]:
            edit_city(col_a)
        elif col_a[0:2] in ["澳洲", "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州", "云南", "陕西", "甘肃", "青海", "台湾", "内蒙古", "内蒙", "广西", "西藏", "宁夏", "新疆", "香港", "澳门", "美国"]:
            pass
        else:
            logger.info(col_a)
        cell_f_num = "F" + str(i)
        wx[cell_f_num] = '是'
        wb.save('2020.xlsx')


def edit_city(col_a):
    logger.info(col_a + "修改好了")


def main():
    read_xls()


if __name__ == "__main__":
    main()
